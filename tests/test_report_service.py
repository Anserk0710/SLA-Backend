from datetime import date, datetime, timezone

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.core.constants import TicketStatus
from app.models.ticket import Ticket
from app.services.report_service import generate_ticket_report_xlsx


def test_generate_ticket_report_xlsx_exports_filtered_ticket_rows(
    db_session: Session,
) -> None:
    matching_ticket = Ticket(
        ticket_code="TCK-REPORT-001",
        full_name="Andi Pratama",
        full_address="Jl. Melati No. 10",
        category="Internet",
        description="Perlu penanganan jaringan.",
        pic_name="Andi",
        phone_number="081200000101",
        internal_status=TicketStatus.ASSIGNED.value,
        public_status="Sedang Dikerjakan",
        sla_deadline=datetime(2020, 1, 1, 0, 0, 0, tzinfo=timezone.utc),
        is_sla_breached=False,
        created_at=datetime(2026, 4, 13, 9, 0, 0, tzinfo=timezone.utc),
    )
    other_ticket = Ticket(
        ticket_code="TCK-REPORT-002",
        full_name="Beni Saputra",
        full_address="Jl. Kenanga No. 11",
        category="CCTV",
        description="Kategori berbeda.",
        pic_name="Beni",
        phone_number="081200000102",
        internal_status=TicketStatus.ASSIGNED.value,
        public_status="Sedang Dikerjakan",
        created_at=datetime(2026, 4, 13, 10, 0, 0, tzinfo=timezone.utc),
    )
    db_session.add_all([matching_ticket, other_ticket])
    db_session.commit()

    output = generate_ticket_report_xlsx(
        db=db_session,
        status="ASSIGNED",
        category="Internet",
        date_from=date(2026, 4, 13),
        date_to=date(2026, 4, 13),
    )

    workbook = load_workbook(output)
    sheet = workbook["Tickets Report"]
    rows = list(sheet.iter_rows(values_only=True))

    db_session.refresh(matching_ticket)

    assert rows[0] == (
        "ID",
        "Ticket Code",
        "Category",
        "Status",
        "SLA Deadline",
        "SLA Breached",
        "Created At",
    )
    assert rows[1] == (
        matching_ticket.id,
        matching_ticket.ticket_code,
        matching_ticket.category,
        matching_ticket.internal_status,
        matching_ticket.sla_deadline.isoformat(),
        "YES",
        matching_ticket.created_at.isoformat(),
    )
    assert len(rows) == 2
    assert matching_ticket.is_sla_breached is True
