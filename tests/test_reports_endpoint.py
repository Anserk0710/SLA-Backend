from datetime import datetime, timezone
from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.api.deps import get_current_user
from app.core.config import settings
from app.core.constants import RoleName, TicketStatus
from app.models.role import Role
from app.models.ticket import Ticket
from app.models.ticket_assignment import TicketAssignment
from app.models.user import User


@pytest.fixture()
def authenticated_admin_client(client: TestClient, db_session: Session) -> TestClient:
    current_user = db_session.scalar(select(User))
    assert current_user is not None

    client.app.dependency_overrides[get_current_user] = lambda: current_user
    try:
        yield client
    finally:
        client.app.dependency_overrides.pop(get_current_user, None)


def test_export_ticket_report_endpoint_returns_filtered_xlsx(
    authenticated_admin_client: TestClient,
    db_session: Session,
) -> None:
    technician_role = Role(name=RoleName.TECHNICIAN.value)
    db_session.add(technician_role)
    db_session.flush()

    technician = User(
        full_name="Teknisi Report",
        email="teknisi.report@example.com",
        hashed_password="hashed-password",
        is_active=True,
        role_id=technician_role.id,
    )
    db_session.add(technician)
    db_session.flush()

    matching_ticket = Ticket(
        ticket_code="TCK-REPORT-ENDPOINT-001",
        full_name="Andi Pratama",
        full_address="Jl. Melati No. 10",
        category="Internet",
        description="Perlu penanganan jaringan.",
        pic_name="Andi",
        phone_number="081200000201",
        internal_status=TicketStatus.ASSIGNED.value,
        public_status="Sedang Dikerjakan",
        sla_deadline=datetime(2020, 1, 1, 0, 0, 0, tzinfo=timezone.utc),
        is_sla_breached=False,
        created_at=datetime(2026, 4, 13, 9, 0, 0, tzinfo=timezone.utc),
    )
    other_ticket = Ticket(
        ticket_code="TCK-REPORT-ENDPOINT-002",
        full_name="Beni Saputra",
        full_address="Jl. Kenanga No. 11",
        category="CCTV",
        description="Kategori berbeda.",
        pic_name="Beni",
        phone_number="081200000202",
        internal_status=TicketStatus.ASSIGNED.value,
        public_status="Sedang Dikerjakan",
        created_at=datetime(2026, 4, 13, 10, 0, 0, tzinfo=timezone.utc),
    )
    db_session.add_all([matching_ticket, other_ticket])
    db_session.flush()

    db_session.add(
        TicketAssignment(
            ticket_id=matching_ticket.id,
            technician_user_id=technician.id,
        )
    )
    db_session.commit()

    response = authenticated_admin_client.get(
        f"{settings.API_V1_STR}/reports/tickets/export",
        params={
            "status": TicketStatus.ASSIGNED.value,
            "category": "Internet",
            "date_from": "2026-04-13",
            "date_to": "2026-04-13",
            "technician_id": technician.id,
        },
    )

    assert response.status_code == 200
    assert (
        response.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert response.headers["content-disposition"].startswith(
        'attachment; filename="ticket-report-'
    )
    assert response.headers["content-disposition"].endswith('.xlsx"')

    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook["Tickets Report"]
    rows = list(sheet.iter_rows(values_only=True))

    db_session.refresh(matching_ticket)

    assert rows[0] == (
        "ID",
        "Ticket Code",
        "Category",
        "Status",
        "SLA Deadline",
        "SLA Breached",
        "Created At",
    )
    assert rows[1] == (
        matching_ticket.id,
        matching_ticket.ticket_code,
        matching_ticket.category,
        matching_ticket.internal_status,
        matching_ticket.sla_deadline.isoformat(),
        "YES",
        matching_ticket.created_at.isoformat(),
    )
    assert len(rows) == 2
